import openpyxl

def makeExcel(datas):
    # 새로운 워크북 생성
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # 엑셀 헤더 셀에 데이터 입력
    worksheet['A1'] = '파일 이름'
    worksheet['B1'] = '파일 사이즈'
    worksheet['C1'] = '파일수정 날짜'
    worksheet['D1'] = '이메일 포함 여부'
    worksheet['E1'] = '주민등록번호 포함 여부'

    # 데이터를 엑셀 시트에 입력
    for row, data in enumerate(datas):
        worksheet.cell(row=row+2, column=1, value=data[0])
        worksheet.cell(row=row+2, column=2, value=data[1])
        worksheet.cell(row=row+2, column=3, value=data[2])
        worksheet.cell(row=row+2, column=4, value=data[3])
        worksheet.cell(row=row+2, column=5, value=data[4])

    # 생성된 워크북을 파일로 저장
    workbook.save('information.xlsx')


